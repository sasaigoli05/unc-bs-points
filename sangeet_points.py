#!/usr/bin/env python3
"""
Sangeet Points Calculator — attendance + tabling + gigs (total-only)

Usage:
  python sangeet_points.py "/path/to/2025-2026 Attendance Sheet.xlsx"

Scoring logic (per your spec):
- Meeting attendance: Each attended meeting is worth 1 / D points,
  where D = (# of subgroups the member is in) + (1 if GBM is included).
  This makes each required weekly meeting share equal (e.g., 2 subgroups + GBM -> D=3 -> 0.33 each).
  ATTENDED = Present, Tardy, Tardy (Excused), Excessive Tardy. Absences (excused/unexcused) are 0.
- Tabling: 1 point per tabling count.
- Gigs: 2 points per gig count.
- Final output sheet "Points Standing" has ONLY two columns: Member | Total_Points

Notes:
- Auto-detects sheet types by name:
    * "GBM" or "General Body" in name -> GBM attendance sheet
    * "Tabling" in name -> Tabling count sheet
    * "Gig" or "Performance" in name -> Gigs count sheet
    * "Points Standing" -> ignored on read (overwritten on write)
    * Everything else (not above) -> subgroup attendance sheets
- Attendance sheets can be WIDE (Name + many date/status columns) or LONG (Name/Member + Status).
- Tabling/Gigs sheets: sums all numeric columns per row (so it works for single "Count" column or per-date 1/0 columns).
- Robust to messy headers and dropdown variants.

If you need to exclude GBM from the attendance denominator, set INCLUDE_GBM_IN_DENOM to False.
"""

import argparse
import math
import re
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ------------------------- Configuration ------------------------- #

# Include GBM as part of the expected weekly meeting count?
INCLUDE_GBM_IN_DENOM = True


# ------------------------- Status normalization ------------------------- #

def _is_nan(x) -> bool:
    try:
        return pd.isna(x)
    except Exception:
        return False

def normalize_status(val) -> str:
    if _is_nan(val):
        return ""
    s = str(val).strip().lower()

    mapping = {
        "present": "present",
        "p": "present",
        "attended": "present",
        "y": "present",
        "yes": "present",
        "here": "present",

        "a": "absent",
        "absent": "absent",
        "n": "absent",
        "no": "absent",

        "excused absence": "excused_absence",
        "excused": "excused_absence",
        "unexcused absence": "unexcused_absence",

        "tardy (excused)": "tardy_excused",
        "tardy(excused)": "tardy_excused",
        "tardy - excused": "tardy_excused",

        "tardy": "tardy",
        "late": "tardy",

        "excessive tardy": "excessive_tardy",
        "excessive tardiness": "excessive_tardy",
    }
    if s in mapping:
        return mapping[s]

    if "excessive" in s and "tard" in s:
        return "excessive_tardy"
    if "tard" in s and "excuse" in s:
        return "tardy_excused"
    if "tard" in s or "late" in s:
        return "tardy"
    if "unexcused" in s and "abs" in s:
        return "unexcused_absence"
    if "excused" in s and "abs" in s:
        return "excused_absence"
    if "abs" in s:
        return "absent"
    if "pres" in s or "attend" in s:
        return "present"
    return s


# ------------------------- Sheet classification ------------------------- #

def classify_sheet(name: str) -> str:
    """
    Returns one of: 'ignore', 'gbm', 'tabling', 'gigs', 'subgroup'
    """
    s = (name or "").strip().lower()
    if "points standing" in s:
        return "ignore"
    if "gbm" in s or "general body" in s:
        return "gbm"
    if "tabling" in s:
        return "tabling"
    if "gig" in s or "performance" in s:
        return "gigs"
    return "subgroup"


# ------------------------- Parsing helpers ------------------------- #

def _stringify_headers(df: pd.DataFrame) -> pd.DataFrame:
    new_cols = []
    seen = set()
    for i, c in enumerate(df.columns):
        if _is_nan(c) or c is None or str(c).strip() == "":
            col = f"col_{i}"
        else:
            col = str(c)
        base = col
        k = 1
        while col in seen:
            col = f"{base}.{k}"
            k += 1
        seen.add(col)
        new_cols.append(col)
    df.columns = new_cols
    return df


def detect_wide_format(df: pd.DataFrame) -> bool:
    """
    Heuristic: first column looks like names, and other columns look like meeting/date/status.
    """
    if df.empty or df.shape[1] < 2:
        return False
    df = _stringify_headers(df.copy())
    first_col = df.columns[0].strip().lower()
    looks_like_name = any(k in first_col for k in ["name", "member", "person", "attendee"])

    date_like_cols = 0
    status_like_cols = 0
    for c in df.columns[1:]:
        cstr = c.strip().lower()
        if re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", cstr) or re.search(r"\d{4}-\d{1,2}-\d{1,2}", cstr) \
           or "week" in cstr or "mtg" in cstr or "rehearsal" in cstr or "meeting" in cstr:
            date_like_cols += 1
        vals = df[c].dropna().astype(str).str.lower().head(20).tolist()
        if vals:
            hits = sum(
                normalize_status(v) in {
                    "present", "absent", "excused_absence", "unexcused_absence",
                    "tardy", "tardy_excused", "excessive_tardy"
                } for v in vals
            )
            if hits / max(1, len(vals)) >= 0.3:
                status_like_cols += 1

    return looks_like_name and (date_like_cols >= 1 or status_like_cols >= 2)


def melt_wide(df: pd.DataFrame) -> pd.DataFrame:
    df = _stringify_headers(df.copy())
    id_col = df.columns[0]
    out = df.melt(id_vars=[id_col], var_name="Meeting", value_name="Status")
    out = out.rename(columns={id_col: "Member"})
    return out


def _guess_member_and_status_cols(df: pd.DataFrame) -> Tuple[str, str]:
    df = _stringify_headers(df.copy())
    cols_lower = {c: c.strip().lower() for c in df.columns}
    member_col: Optional[str] = None
    status_col: Optional[str] = None
    for c, lc in cols_lower.items():
        if member_col is None and any(k in lc for k in ["member", "name", "person", "attendee"]):
            member_col = c
        if status_col is None and any(k in lc for k in ["status", "attendance"]):
            status_col = c
    if member_col is None:
        member_col = df.columns[0]
    if status_col is None:
        best_col, best_score = None, -1.0
        for c in df.columns:
            if c == member_col:
                continue
            vals = df[c].dropna().astype(str).str.lower().tolist()
            if not vals:
                continue
            hits = sum(
                normalize_status(v) in {
                    "present", "absent", "excused_absence", "unexcused_absence",
                    "tardy", "tardy_excused", "excessive_tardy"
                } for v in vals
            )
            score = hits / max(1, len(vals))
            if score > best_score:
                best_score, best_col = score, c
        status_col = best_col if best_col is not None else df.columns[-1]
    return member_col, status_col


def normalize_attendance_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["Member", "Status"])
    df = df.dropna(how="all").dropna(axis=1, how="all")
    if df.empty:
        return pd.DataFrame(columns=["Member", "Status"])

    if detect_wide_format(df):
        out = melt_wide(df)
    else:
        mcol, scol = _guess_member_and_status_cols(df)
        df = _stringify_headers(df)
        if mcol not in df.columns:
            mcol = df.columns[0]
        if scol not in df.columns:
            scol = df.columns[-1]
        out = df.rename(columns={mcol: "Member", scol: "Status"})[["Member", "Status"]]

    out["Member"] = out["Member"].astype(str).str.strip()
    out["Status"] = out["Status"].apply(normalize_status)
    out = out[~out["Member"].isna() & (out["Member"].str.strip() != "")]
    return out


# ------------------------- Counting helpers ------------------------- #

def status_counts(df: pd.DataFrame, member: str) -> Dict[str, int]:
    out = {
        "present": 0,
        "absent": 0,
        "excused_absence": 0,
        "unexcused_absence": 0,
        "tardy": 0,
        "tardy_excused": 0,
        "excessive_tardy": 0,
    }
    if df.empty:
        return out
    sub = df[df["Member"] == member]
    for s in sub["Status"].tolist():
        ns = normalize_status(s)
        if ns in out:
            out[ns] += 1
        else:
            # treat unknown non-empty non-absence as attended
            if ns and ns not in {"absent", "excused_absence", "unexcused_absence", ""}:
                out["present"] += 1
    return out


def sum_numeric_row_values(df: pd.DataFrame, member_col_guess: Optional[str] = None) -> pd.DataFrame:
    """
    For Tabling/Gigs sheets: pick the best member column, then sum all numeric-ish columns per row.
    Returns DataFrame with columns: Member, Count
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=["Member", "Count"])
    df = _stringify_headers(df.dropna(how="all").dropna(axis=1, how="all"))
    if df.empty:
        return pd.DataFrame(columns=["Member", "Count"])

    # Guess member column
    if member_col_guess is None:
        member_col, _ = _guess_member_and_status_cols(df)
    else:
        member_col = member_col_guess if member_col_guess in df.columns else df.columns[0]

    # Convert all non-member columns to numeric (coerce), sum across rows
    num_df = df.drop(columns=[member_col], errors="ignore").apply(pd.to_numeric, errors="coerce")
    counts = num_df.sum(axis=1, skipna=True).fillna(0.0)
    out = pd.DataFrame({"Member": df[member_col].astype(str).str.strip(), "Count": counts})
    out = out[~out["Member"].isna() & (out["Member"].str.strip() != "")]
    return out


# ------------------------- Main compute ------------------------- #

def compute_points(workbook_path: str) -> None:
    wb = load_workbook(workbook_path)
    sheet_names = wb.sheetnames

    gbm_frames: List[pd.DataFrame] = []
    subgroup_frames_by_name: Dict[str, pd.DataFrame] = {}
    tabling_frames: List[pd.DataFrame] = []
    gigs_frames: List[pd.DataFrame] = []

    # First pass: read & classify
    for sname in sheet_names:
        kind = classify_sheet(sname)
        if kind == "ignore":
            continue

        ws = wb[sname]
        values = list(ws.values)
        if not values:
            continue

        header = list(values[0]) if values else []
        data = values[1:] if len(values) > 1 else []
        df = pd.DataFrame(data, columns=header).dropna(how="all")
        if df.empty:
            continue

        if kind == "tabling":
            tabling_frames.append(sum_numeric_row_values(df))
            continue

        if kind == "gigs":
            gigs_frames.append(sum_numeric_row_values(df))
            continue

        # Attendance sheets
        att = normalize_attendance_frame(df)
        if kind == "gbm":
            att["Sheet"] = sname
            gbm_frames.append(att)
        else:  # subgroup
            att["Sheet"] = sname
            subgroup_frames_by_name[sname] = att

    gbm_att = pd.concat(gbm_frames, ignore_index=True) if gbm_frames else pd.DataFrame(columns=["Member", "Status", "Sheet"])
    subgroup_att = pd.concat(list(subgroup_frames_by_name.values()), ignore_index=True) if subgroup_frames_by_name else pd.DataFrame(columns=["Member", "Status", "Sheet"])
    tabling_counts = pd.concat(tabling_frames, ignore_index=True) if tabling_frames else pd.DataFrame(columns=["Member", "Count"])
    gigs_counts = pd.concat(gigs_frames, ignore_index=True) if gigs_frames else pd.DataFrame(columns=["Member", "Count"])

    # Member universe: anyone appearing anywhere
    members = set()
    for frame in (gbm_att, subgroup_att, tabling_counts, gigs_counts):
        if not frame.empty and "Member" in frame.columns:
            members.update([m for m in frame["Member"].tolist() if isinstance(m, str) and m.strip()])

    # Infer number of subgroups per member based on distinct subgroup sheets they appear on
    member_to_subgroups: Dict[str, int] = {m: 0 for m in members}
    for sname, att in subgroup_frames_by_name.items():
        if att.empty:
            continue
        for m in set(att["Member"].unique().tolist()):
            member_to_subgroups[m] = member_to_subgroups.get(m, 0) + 1

    gbm_exists = not gbm_att.empty
    include_gbm = (INCLUDE_GBM_IN_DENOM and gbm_exists)

    # Compute total points
    rows = []
    for m in sorted(members):
        n_subgroups = member_to_subgroups.get(m, 0)
        denom = n_subgroups + (1 if include_gbm else 0)
        per_meeting_points = 0.0 if denom == 0 else 1.0 / float(denom)

        # Attendance points: add per-meeting for each attended meeting across GBM (if included) + subgroups
        total_attendance_points = 0.0

        # GBM attended meetings
        if include_gbm and not gbm_att.empty:
            gcounts = status_counts(gbm_att, m)
            gbm_attended = gcounts["present"] + gcounts["tardy"] + gcounts["tardy_excused"] + gcounts["excessive_tardy"]
            total_attendance_points += gbm_attended * per_meeting_points

        # Subgroup attended meetings
        if not subgroup_att.empty:
            scounts = status_counts(subgroup_att, m)
            sg_attended = scounts["present"] + scounts["tardy"] + scounts["tardy_excused"] + scounts["excessive_tardy"]
            total_attendance_points += sg_attended * per_meeting_points

        # Tabling points: 1 per count
        tab_points = 0.0
        if not tabling_counts.empty:
            tab_points = float(tabling_counts.loc[tabling_counts["Member"] == m, "Count"].sum()) * 1.0

        # Gigs points: 2 per count
        gig_points = 0.0
        if not gigs_counts.empty:
            gig_points = float(gigs_counts.loc[gigs_counts["Member"] == m, "Count"].sum()) * 2.0

        total_points = total_attendance_points + tab_points + gig_points
        rows.append({"Member": m, "Total_Points": total_points})

    summary = pd.DataFrame(rows).sort_values(["Total_Points", "Member"], ascending=[False, True])

    # Overwrite/create "Points Standing" as last sheet with ONLY two columns
    to_delete = [s for s in wb.sheetnames if s.strip().lower() == "points standing"]
    for s in to_delete:
        del wb[s]
    ws = wb.create_sheet("Points Standing")
    for r in dataframe_to_rows(summary[["Member", "Total_Points"]], index=False, header=True):
        ws.append(r)

    wb.save(workbook_path)


# ------------------------- CLI ------------------------- #

def main():
    parser = argparse.ArgumentParser(description="Compute total Sangeet points (attendance + tabling + gigs) and write to 'Points Standing'.")
    parser.add_argument("xlsx_path", help="Path to the attendance .xlsx workbook")
    args = parser.parse_args()
    compute_points(args.xlsx_path)


if __name__ == "__main__":
    main()
